#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
批量重命名工具
功能：
1. 读取Excel表（原名称、新名称两列）
2. 模式A: 扫描文件夹（含子文件夹）中的所有WAV文件，根据Excel表重命名
3. 模式B: 选择父级文件夹，读取第一层子文件夹名称，匹配Excel表重命名文件夹
4. 模式C: 输入指定字符串，删除WAV文件名中的该字符串并重命名
5. 在原Excel表上标记重命名结果（增量重命名）
6. 支持撤销当前批次的重命名操作
7. 预览区精简展示，支持交互修正/删除
8. 日志区默认收起，可展开
"""

import os
import sys
import uuid
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("请先安装openpyxl: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────
#  辅助：可折叠的日志区
# ─────────────────────────────────────────────
class CollapsibleLog(ttk.Frame):
    """可折叠的日志区组件"""

    def __init__(self, parent, default_open=False, **kwargs):
        super().__init__(parent, **kwargs)
        self._open = tk.BooleanVar(value=default_open)

        # 标题行
        header = ttk.Frame(self)
        header.pack(fill=tk.X)

        self._toggle_btn = ttk.Button(header, text="▶ 操作日志（展开）",
                                      command=self._toggle, width=20)
        self._toggle_btn.pack(side=tk.LEFT)
        ttk.Button(header, text="清除", command=self._clear, width=6).pack(side=tk.LEFT, padx=4)

        # 日志内容区
        self._body = ttk.Frame(self)
        self.log_text = scrolledtext.ScrolledText(self._body, wrap=tk.WORD, height=8)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.tag_config('info', foreground='#333333')
        self.log_text.tag_config('success', foreground='#008000')
        self.log_text.tag_config('warning', foreground='#cc6600')
        self.log_text.tag_config('error', foreground='#cc0000')

        if default_open:
            self._body.pack(fill=tk.BOTH, expand=True)
            self._toggle_btn.config(text="▼ 操作日志（收起）")

    def _toggle(self):
        if self._open.get():
            self._body.pack_forget()
            self._open.set(False)
            self._toggle_btn.config(text="▶ 操作日志（展开）")
        else:
            self._body.pack(fill=tk.BOTH, expand=True)
            self._open.set(True)
            self._toggle_btn.config(text="▼ 操作日志（收起）")

    def _clear(self):
        self.log_text.delete(1.0, tk.END)

    def write(self, message, level='info'):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", level)
        self.log_text.see(tk.END)


# ─────────────────────────────────────────────
#  预览编辑器：精简可交互的重命名列表
#  执行后可回写每行状态（成功/失败/跳过）
# ─────────────────────────────────────────────
class PreviewEditor(ttk.Frame):
    """
    预览编辑区
    rows: list of dict  {"old": str, "new": str, ...任意附加字段}
    - 预览阶段：可修改新名称、点击 ✕ 删除行
    - 执行后：通过 mark_result(idx, status, reason, path) 更新每行状态
              📂 按钮可打开文件所在目录
    """

    COL_IDX   = 0
    COL_OLD   = 1
    COL_NEW   = 2
    COL_ACT   = 3
    COL_STATE = 4

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self._rows: List[Dict] = []
        self._entry_vars: List[tk.StringVar] = []
        self._deleted: List[bool] = []
        self._row_frames = []
        self._entries: List[ttk.Entry] = []
        self._del_btns: List[ttk.Button] = []
        self._result_labels: List[Optional[tk.Label]] = []

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        # ── 表头
        hdr = tk.Frame(self, bg='#4472C4')
        hdr.grid(row=0, column=0, sticky='ew')
        hdr.columnconfigure(self.COL_OLD, weight=1)
        hdr.columnconfigure(self.COL_NEW, weight=1)
        hdr.columnconfigure(self.COL_STATE, weight=1)
        for col, (text, w) in enumerate([
            ("#", 4), ("原名称", 20), ("新名称（可编辑）", 20), ("操作", 5), ("执行状态", 20)
        ]):
            tk.Label(hdr, text=text, width=w, anchor='w',
                     bg='#4472C4', fg='white',
                     font=('Arial', 9, 'bold')).grid(
                row=0, column=col, padx=3, pady=3, sticky='w')

        # ── 滚动容器
        sc = tk.Frame(self, bg='#cccccc')
        sc.grid(row=1, column=0, sticky='nsew')
        sc.columnconfigure(0, weight=1)
        sc.rowconfigure(0, weight=1)

        self._canvas = tk.Canvas(sc, bg='white', highlightthickness=0, bd=0)
        vbar = ttk.Scrollbar(sc, orient=tk.VERTICAL, command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vbar.set)
        self._canvas.grid(row=0, column=0, sticky='nsew')
        vbar.grid(row=0, column=1, sticky='ns')

        self._inner = tk.Frame(self._canvas, bg='white')
        self._inner_id = self._canvas.create_window((0, 0), window=self._inner, anchor='nw')

        self._canvas.bind('<Configure>', self._on_canvas_resize)
        self._inner.bind('<Configure>', self._on_inner_resize)
        for w in (self._canvas, self._inner):
            w.bind('<MouseWheel>',
                   lambda e: self._canvas.yview_scroll(-1 * (e.delta // 120), 'units'))

        self._inner.columnconfigure(self.COL_OLD, weight=1)
        self._inner.columnconfigure(self.COL_NEW, weight=1)
        self._inner.columnconfigure(self.COL_STATE, weight=1)

        tk.Label(self._inner, text="（暂无数据，请先点击【预览重命名】）",
                 fg='#999999', bg='white', font=('Arial', 10)).grid(
            row=0, column=0, columnspan=5, pady=20)

    def _on_canvas_resize(self, event):
        self._canvas.itemconfig(self._inner_id, width=event.width)

    def _on_inner_resize(self, event):
        self._canvas.configure(scrollregion=self._canvas.bbox('all'))

    def load_rows(self, rows: List[Dict]):
        """加载预览数据（必须在主线程调用）"""
        for w in self._inner.winfo_children():
            w.destroy()
        self._rows = rows
        self._entry_vars = []
        self._deleted = [False] * len(rows)
        self._row_frames = []
        self._entries = []
        self._del_btns = []
        self._result_labels = [None] * len(rows)

        if not rows:
            tk.Label(self._inner, text="（暂无匹配项）",
                     fg='#999999', bg='white', font=('Arial', 10)).grid(
                row=0, column=0, columnspan=5, pady=20)
            self._canvas.configure(scrollregion=self._canvas.bbox('all'))
            return

        self._inner.columnconfigure(self.COL_OLD, weight=1)
        self._inner.columnconfigure(self.COL_NEW, weight=1)
        self._inner.columnconfigure(self.COL_STATE, weight=1)

        for idx, row in enumerate(rows):
            var = tk.StringVar(value=row.get('new', ''))
            self._entry_vars.append(var)
            self._build_row(idx, row, var)

        self._inner.update_idletasks()
        self._canvas.configure(scrollregion=self._canvas.bbox('all'))
        self._canvas.yview_moveto(0)

    def _build_row(self, idx: int, row: Dict, var: tk.StringVar):
        bg = '#f0f4ff' if idx % 2 == 0 else '#ffffff'

        tk.Label(self._inner, text=str(idx + 1), width=4, anchor='e',
                 bg=bg, fg='#666666', font=('Arial', 9)).grid(
            row=idx, column=self.COL_IDX, padx=(4, 2), pady=1, sticky='e')

        lbl = tk.Label(self._inner, text=row.get('old', ''), anchor='w',
                       bg=bg, fg='#333333', font=('Arial', 9))
        lbl.grid(row=idx, column=self.COL_OLD, padx=4, pady=1, sticky='ew')

        entry = ttk.Entry(self._inner, textvariable=var)
        entry.grid(row=idx, column=self.COL_NEW, padx=4, pady=1, sticky='ew')
        entry.bind('<MouseWheel>',
                   lambda e: self._canvas.yview_scroll(-1 * (e.delta // 120), 'units'))

        btn = ttk.Button(self._inner, text="✕", width=3,
                         command=lambda i=idx: self._toggle_delete(i))
        btn.grid(row=idx, column=self.COL_ACT, padx=4, pady=1)

        state_lbl = tk.Label(self._inner, text="", anchor='w',
                             bg=bg, fg='#555', font=('Arial', 9))
        state_lbl.grid(row=idx, column=self.COL_STATE, padx=4, pady=1, sticky='ew')
        self._result_labels[idx] = state_lbl

        self._row_frames.append((lbl, entry, btn, bg))
        self._entries.append(entry)
        self._del_btns.append(btn)

    def _toggle_delete(self, idx: int):
        if self._deleted[idx]:
            self._deleted[idx] = False
            _, entry, btn, bg = self._row_frames[idx]
            entry.configure(state='normal')
            btn.configure(text='✕')
            self._set_row_bg(idx, bg)
        else:
            self._deleted[idx] = True
            _, entry, btn, _ = self._row_frames[idx]
            entry.configure(state='disabled')
            btn.configure(text='↩')
            self._set_row_bg(idx, '#ffdddd')

    def _set_row_bg(self, idx: int, color: str):
        for col in range(5):
            ws = self._inner.grid_slaves(row=idx, column=col)
            if ws:
                try:
                    ws[0].configure(bg=color)
                except Exception:
                    pass

    def mark_result(self, idx: int, status: str, reason: str = '',
                    result_path: str = ''):
        """
        执行后更新某行状态（在主线程调用）。
        status: '成功' | '失败' | '跳过'
        reason: 失败/跳过原因说明
        result_path: 成功时=新路径；失败时=原路径（用于定位）
        """
        if idx >= len(self._result_labels) or self._result_labels[idx] is None:
            return
        state_lbl = self._result_labels[idx]
        _, entry, btn, bg = self._row_frames[idx]

        if status == '成功':
            row_bg = '#d6f0d6'
            state_lbl.configure(text='✅ 成功', fg='#006100',
                                 bg=row_bg, font=('Arial', 9, 'bold'))
            entry.configure(state='disabled')
            btn.configure(text='📂', width=3,
                          command=lambda p=result_path: self._open_location(p))
        elif status == '失败':
            row_bg = '#ffd6d6'
            tip = f'❌ {reason}' if reason else '❌ 失败'
            state_lbl.configure(text=tip, fg='#9c0006', bg=row_bg, font=('Arial', 9))
            btn.configure(text='📂', width=3,
                          command=lambda p=result_path: self._open_location(p))
        elif status == '跳过':
            row_bg = '#fff3cd'
            tip = f'⏭ {reason}' if reason else '⏭ 跳过'
            state_lbl.configure(text=tip, fg='#856404', bg=row_bg, font=('Arial', 9))
        else:
            row_bg = bg

        self._set_row_bg(idx, row_bg)

    def _open_location(self, path: str):
        """在资源管理器中打开文件/文件夹所在目录"""
        if not path:
            return
        p = Path(path)
        try:
            target = str(p.parent) if p.is_file() else str(p if p.exists() else p.parent)
            if Path(target).exists():
                os.startfile(target)
            else:
                messagebox.showwarning("提示", f"路径不存在:\n{path}")
        except Exception as e:
            messagebox.showerror("错误", f"打开路径失败:\n{e}")

    def get_rows(self) -> List[Dict]:
        """返回未删除且新名称非空的行，每行附带 _preview_idx 字段（对应预览区行号）"""
        result = []
        for idx, row in enumerate(self._rows):
            if self._deleted[idx]:
                continue
            new_val = self._entry_vars[idx].get().strip()
            if not new_val:
                continue
            merged = dict(row)
            merged['new'] = new_val
            merged['_preview_idx'] = idx   # 记录预览行号，执行后回写状态用
            result.append(merged)
        return result

    def is_empty(self) -> bool:
        return len(self._rows) == 0

    def count(self) -> int:
        return sum(1 for d in self._deleted if not d)


# ─────────────────────────────────────────────
#  主工具类
# ─────────────────────────────────────────────
class WavRenamerApp:
    """重命名工具主界面"""

    def __init__(self, root):
        self.root = root
        self.root.title("批量重命名工具")
        self.root.geometry("1020x800")
        self.root.minsize(900, 650)

        # 公共变量
        self.excel_path = tk.StringVar()
        self.folder_path = tk.StringVar()       # WAV文件夹 / 父级文件夹
        self.rename_mode = tk.StringVar(value="wav")  # "wav" | "folder"

        # 删除字符串模式
        self.strip_folder_path = tk.StringVar()
        self.strip_string = tk.StringVar()

        # 运行时数据
        self.rename_results: List[Dict] = []
        self.undo_history: List[Dict] = []

        self._setup_ui()

    # ──────────────────────── UI 搭建 ────────────────────────
    def _setup_ui(self):
        """搭建整体 UI"""
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        nb = ttk.Notebook(self.root)
        nb.grid(row=0, column=0, sticky='nsew', padx=6, pady=6)

        # Tab 1: 读表重命名（WAV / 文件夹双模式）
        self._tab_rename = ttk.Frame(nb)
        nb.add(self._tab_rename, text=" 读表重命名 ")
        self._build_rename_tab(self._tab_rename)

        # Tab 2: 删除字符串重命名 WAV
        self._tab_strip = ttk.Frame(nb)
        nb.add(self._tab_strip, text=" 删除字符串重命名WAV ")
        self._build_strip_tab(self._tab_strip)

    # ──────── Tab1: 读表重命名 ────────
    def _build_rename_tab(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(4, weight=1)   # 预览区
        parent.rowconfigure(5, weight=0)   # 日志区

        row_i = 0
        pad = {'padx': 8, 'pady': 4}

        # ── 模式切换
        mode_frame = ttk.LabelFrame(parent, text="重命名模式", padding=6)
        mode_frame.grid(row=row_i, column=0, sticky='ew', **pad)
        ttk.Radiobutton(mode_frame, text="WAV文件重命名（递归扫描WAV文件，精确匹配）",
                        variable=self.rename_mode, value="wav",
                        command=self._on_mode_change).pack(side=tk.LEFT, padx=8)
        ttk.Radiobutton(mode_frame, text="文件夹名称重命名（扫描第一层子文件夹，包含匹配）",
                        variable=self.rename_mode, value="folder",
                        command=self._on_mode_change).pack(side=tk.LEFT, padx=8)
        row_i += 1

        # ── Excel 选择
        excel_frame = ttk.LabelFrame(parent, text="Excel改名表", padding=6)
        excel_frame.grid(row=row_i, column=0, sticky='ew', **pad)
        excel_frame.columnconfigure(0, weight=1)
        ttk.Entry(excel_frame, textvariable=self.excel_path, state='readonly').grid(
            row=0, column=0, sticky='ew', padx=(0, 4))
        ttk.Button(excel_frame, text="选择Excel", command=self.select_excel, width=10).grid(
            row=0, column=1)
        ttk.Button(excel_frame, text="打开", command=self.open_excel, width=6).grid(
            row=0, column=2, padx=(4, 0))

        # 列设置（只读原名列、新名列、起始行，不再使用状态列/路径列）
        col_sub = ttk.Frame(excel_frame)
        col_sub.grid(row=1, column=0, columnspan=3, sticky='w', pady=(4, 0))
        for label, attr, default in [
            ("原名称列", "old_col", "A"),
            ("新名称列", "new_col", "B"),
            ("起始行", "start_row", "2"),
        ]:
            ttk.Label(col_sub, text=f"{label}:").pack(side=tk.LEFT, padx=(8, 2))
            e = ttk.Entry(col_sub, width=4)
            e.insert(0, default)
            e.pack(side=tk.LEFT, padx=(0, 4))
            setattr(self, attr, e)
        row_i += 1

        # ── 文件夹选择
        folder_lf = ttk.LabelFrame(parent, text="扫描目标文件夹", padding=6)
        folder_lf.grid(row=row_i, column=0, sticky='ew', **pad)
        folder_lf.columnconfigure(0, weight=1)
        self._folder_label = ttk.Label(folder_lf, text="WAV文件夹:")
        self._folder_label.grid(row=0, column=0, sticky='w', columnspan=3)

        folder_inner = ttk.Frame(folder_lf)
        folder_inner.grid(row=1, column=0, columnspan=3, sticky='ew')
        folder_inner.columnconfigure(0, weight=1)
        ttk.Entry(folder_inner, textvariable=self.folder_path, state='readonly').grid(
            row=0, column=0, sticky='ew', padx=(0, 4))
        ttk.Button(folder_inner, text="选择文件夹", command=self.select_folder, width=10).grid(
            row=0, column=1)
        ttk.Button(folder_inner, text="打开", command=self.open_folder, width=6).grid(
            row=0, column=2, padx=(4, 0))
        row_i += 1

        # ── 操作按钮
        btn_frame = ttk.Frame(parent)
        btn_frame.grid(row=row_i, column=0, pady=8)
        ttk.Button(btn_frame, text="🔍 预览重命名", command=self.preview_rename,
                   width=15).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="▶ 执行重命名", command=self.execute_rename,
                   width=15).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="↩ 撤销重命名", command=self.undo_rename,
                   width=15).grid(row=0, column=2, padx=5)

        # 进度条（内嵌到按钮行旁）
        self.progress = ttk.Progressbar(btn_frame, mode='determinate', length=200)
        self.progress.grid(row=0, column=3, padx=10)
        row_i += 1

        # ── 预览区（精简交互列表）
        preview_lf = ttk.LabelFrame(parent, text="预览区（可修改新名称 / 点击 ✕ 删除该行）", padding=6)
        preview_lf.grid(row=row_i, column=0, sticky='nsew', **pad)
        preview_lf.columnconfigure(0, weight=1)
        preview_lf.rowconfigure(0, weight=1)
        self.preview_editor = PreviewEditor(preview_lf)
        self.preview_editor.pack(fill=tk.BOTH, expand=True)
        row_i += 1

        # ── 日志区（默认收起）
        self.log_panel = CollapsibleLog(parent, default_open=False)
        self.log_panel.grid(row=row_i, column=0, sticky='nsew', **pad)
        row_i += 1

        # 欢迎日志
        self._log("欢迎使用批量重命名工具 ！", 'success')
        self._log("模式A-WAV文件重命名：选择Excel + WAV文件夹，预览后执行。", 'info')
        self._log("模式B-文件夹重命名：选择Excel + 父级文件夹，自动匹配子文件夹名称。", 'info')
        self._log("Tab「删除字符串重命名WAV」：批量去除文件名中的指定字符串。", 'info')

    # ──────── Tab2: 删除字符串重命名 ────────
    def _build_strip_tab(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(3, weight=1)   # 预览区
        parent.rowconfigure(4, weight=0)   # 日志区

        pad = {'padx': 8, 'pady': 4}
        row_i = 0

        # ── 参数区
        cfg_frame = ttk.LabelFrame(parent, text="配置", padding=8)
        cfg_frame.grid(row=row_i, column=0, sticky='ew', **pad)
        cfg_frame.columnconfigure(1, weight=1)

        ttk.Label(cfg_frame, text="目标文件夹:").grid(row=0, column=0, sticky='w', pady=4)
        folder_inner = ttk.Frame(cfg_frame)
        folder_inner.grid(row=0, column=1, sticky='ew', pady=4)
        folder_inner.columnconfigure(0, weight=1)
        ttk.Entry(folder_inner, textvariable=self.strip_folder_path, state='readonly').grid(
            row=0, column=0, sticky='ew', padx=(0, 4))
        ttk.Button(folder_inner, text="选择文件夹", width=10,
                   command=self._select_strip_folder).grid(row=0, column=1)
        ttk.Button(folder_inner, text="打开", width=6,
                   command=self._open_strip_folder).grid(row=0, column=2, padx=(4, 0))

        ttk.Label(cfg_frame, text="要删除的字符串:").grid(row=1, column=0, sticky='w', pady=4)
        strip_inner = ttk.Frame(cfg_frame)
        strip_inner.grid(row=1, column=1, sticky='ew', pady=4)
        strip_inner.columnconfigure(0, weight=1)
        self._strip_entry = ttk.Entry(strip_inner, textvariable=self.strip_string, width=40)
        self._strip_entry.grid(row=0, column=0, sticky='ew', padx=(0, 8))
        ttk.Label(strip_inner, text="（从WAV文件名中删除此字符串）",
                  foreground='#666666').grid(row=0, column=1, sticky='w')
        row_i += 1

        # ── 按钮
        btn_frame = ttk.Frame(parent)
        btn_frame.grid(row=row_i, column=0, pady=8)
        ttk.Button(btn_frame, text="🔍 预览重命名", command=self._strip_preview,
                   width=15).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="▶ 执行重命名", command=self._strip_execute,
                   width=15).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="↩ 撤销重命名", command=self._strip_undo,
                   width=15).grid(row=0, column=2, padx=5)
        self.strip_progress = ttk.Progressbar(btn_frame, mode='determinate', length=200)
        self.strip_progress.grid(row=0, column=3, padx=10)
        row_i += 1

        # ── 预览区
        preview_lf = ttk.LabelFrame(parent,
                                    text="预览区（可修改新名称 / 点击 ✕ 删除该行）", padding=6)
        preview_lf.grid(row=row_i, column=0, sticky='nsew', **pad)
        preview_lf.columnconfigure(0, weight=1)
        preview_lf.rowconfigure(0, weight=1)
        self.strip_preview_editor = PreviewEditor(preview_lf)
        self.strip_preview_editor.pack(fill=tk.BOTH, expand=True)
        row_i += 1

        # ── 日志区
        self.strip_log_panel = CollapsibleLog(parent, default_open=False)
        self.strip_log_panel.grid(row=row_i, column=0, sticky='nsew', **pad)

    # ──────────────────────── 日志 ────────────────────────
    def _log(self, msg, level='info'):
        """写入 Tab1 日志区"""
        self.log_panel.write(msg, level)
        self.root.update_idletasks()

    def _strip_log(self, msg, level='info'):
        """写入 Tab2 日志区"""
        self.strip_log_panel.write(msg, level)
        self.root.update_idletasks()

    # ──────────────────────── 模式切换回调 ────────────────────────
    def _on_mode_change(self):
        mode = self.rename_mode.get()
        if mode == "wav":
            self._folder_label.config(text="WAV文件夹（递归扫描）:")
        else:
            self._folder_label.config(text="父级文件夹（扫描第一层子文件夹名称）:")
        # 清空预览
        self.preview_editor.load_rows([])

    # ──────────────────────── 文件/文件夹选择 ────────────────────────
    def select_excel(self):
        filename = filedialog.askopenfilename(
            title="选择Excel改名表",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.excel_path.set(filename)
            self._log(f"已选择Excel文件: {filename}", 'success')
            self._check_and_init_excel_columns()

    def select_folder(self):
        folder = filedialog.askdirectory(title="选择目标文件夹")
        if folder:
            self.folder_path.set(folder)
            self._log(f"已选择文件夹: {folder}", 'success')

    def open_excel(self):
        if self.excel_path.get():
            os.startfile(self.excel_path.get())
        else:
            messagebox.showwarning("提示", "请先选择Excel文件！")

    def open_folder(self):
        if self.folder_path.get():
            os.startfile(self.folder_path.get())
        else:
            messagebox.showwarning("提示", "请先选择文件夹！")

    def _select_strip_folder(self):
        folder = filedialog.askdirectory(title="选择WAV文件夹")
        if folder:
            self.strip_folder_path.set(folder)
            self._strip_log(f"已选择文件夹: {folder}", 'success')

    def _open_strip_folder(self):
        if self.strip_folder_path.get():
            os.startfile(self.strip_folder_path.get())
        else:
            messagebox.showwarning("提示", "请先选择文件夹！")

    # ──────────────────────── Excel 操作 ────────────────────────
    def _check_and_init_excel_columns(self):
        """不再写入Excel状态列，此方法保留但不执行任何操作"""
        pass

    def read_excel_mapping(self) -> Dict[str, Dict]:
        """读取Excel重命名映射（读取所有有原名+新名的行，不检查状态列）"""
        excel_file = self.excel_path.get()
        old_col = self.old_col.get().strip().upper()
        new_col = self.new_col.get().strip().upper()
        start_row_val = self.start_row.get().strip()
        try:
            start_row = int(start_row_val)
        except ValueError:
            self._log(f"起始行设置无效: '{start_row_val}'，使用默认值 2", 'warning')
            start_row = 2

        self._log(f"读取Excel列设置: 原名={old_col}, 新名={new_col}, 起始行={start_row}", 'info')
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            self._log(f"Excel共 {ws.max_row} 行 × {ws.max_column} 列，活动Sheet: {ws.title}", 'info')

            # 打印第1行内容帮助确认列
            row1_vals = []
            for c in range(1, min(ws.max_column + 1, 8)):
                cell = ws.cell(row=1, column=c)
                row1_vals.append(f"{cell.column_letter}={repr(cell.value)}")
            self._log(f"第1行内容: {', '.join(row1_vals)}", 'info')

            rename_map = {}
            empty_count = 0
            for row in range(start_row, ws.max_row + 1):
                old_name = ws[f"{old_col}{row}"].value
                new_name = ws[f"{new_col}{row}"].value
                if old_name and new_name:
                    old_name = str(old_name).strip()
                    new_name = str(new_name).strip()
                    if old_name.lower().endswith('.wav'):
                        old_name = old_name[:-4]
                    if new_name.lower().endswith('.wav'):
                        new_name = new_name[:-4]
                    if old_name and new_name:
                        rename_map[old_name] = {'new_name': new_name, 'row': row}
                    empty_count = 0   # 有数据行，重置空行计数
                else:
                    if old_name is None and new_name is None:
                        empty_count += 1

                # 当连续超过5行都是空行时，停止读取
                if empty_count >= 5:
                    break

            wb.close()
            return rename_map
        except Exception as e:
            self._log(f"读取Excel失败: {e}", 'error')
            messagebox.showerror("错误", f"读取Excel失败:\n{e}")
            return {}

    def update_excel_with_results(self):
        """不再回写Excel状态，此方法保留但不执行任何操作"""
        pass

    # ──────────────────────── WAV扫描 ────────────────────────
    def get_all_wav_files(self, folder_path: str) -> Dict[str, str]:
        """递归获取所有WAV文件, 返回 {stem: full_path}"""
        wav_files = {}
        folder = Path(folder_path)
        for wav_file in folder.rglob("*.wav"):
            stem = wav_file.stem
            if stem in wav_files:
                self._log(f"警告: 重复文件名 '{stem}'，后者覆盖前者", 'warning')
            wav_files[stem] = str(wav_file)
        return wav_files

    # ──────────────────────── 文件夹扫描（模式B）────────────────────────
    def get_first_level_subfolders(self, parent_path: str) -> Dict[str, str]:
        """获取父级文件夹的第一层子文件夹, 返回 {folder_name: full_path}"""
        result = {}
        parent = Path(parent_path)
        for item in parent.iterdir():
            if item.is_dir():
                result[item.name] = str(item)
        return result

    # ──────────────────────── 依赖分析 ────────────────────────
    def analyze_rename_dependencies(self, rename_map: Dict[str, Dict],
                                    existing_names: Dict[str, str]) -> Tuple[List[List[str]], set]:
        simple_map = {k: v['new_name'] for k, v in rename_map.items()}
        files_in_conflicts = set()
        for old_name, info in rename_map.items():
            new_name = info['new_name']
            if new_name == old_name:
                continue
            if new_name in existing_names and new_name in rename_map:
                files_in_conflicts.add(old_name)
        cycles = []
        for old_name in files_in_conflicts:
            cycle = self._find_cycle(old_name, simple_map, set())
            if cycle and cycle not in cycles:
                cycles.append(cycle)
        if cycles:
            self._log(f"检测到 {len(cycles)} 个名称交换循环:", 'warning')
            for cycle in cycles:
                self._log(f"  循环: {' -> '.join(cycle)} -> {cycle[0]}", 'warning')
        if files_in_conflicts:
            self._log(f"检测到 {len(files_in_conflicts)} 个命名冲突，将使用临时名称中转", 'warning')
        return cycles, files_in_conflicts

    def _find_cycle(self, start: str, rename_map: Dict[str, str],
                    visited: set, path: List[str] = None) -> Optional[List[str]]:
        if path is None:
            path = []
        if start in visited:
            if start in path:
                return path[path.index(start):]
            return None
        visited.add(start)
        path.append(start)
        if start in rename_map:
            return self._find_cycle(rename_map[start], rename_map, visited, path)
        return None

    # ──────────────────────── 预览重命名（Tab1）────────────────────────
    def preview_rename(self):
        if not self.excel_path.get():
            messagebox.showwarning("提示", "请先选择Excel文件！")
            return
        if not self.folder_path.get():
            messagebox.showwarning("提示", "请先选择目标文件夹！")
            return

        def _run():
            mode = self.rename_mode.get()
            self._log(f"{'='*50}", 'info')
            self._log(f"开始预览（{'WAV文件模式' if mode == 'wav' else '文件夹名称模式'}）...", 'info')

            rename_map = self.read_excel_mapping()
            if not rename_map:
                self._log("Excel表中没有待重命名的项目！", 'warning')
                return
            self._log(f"从Excel读取 {len(rename_map)} 条待重命名记录", 'success')

            # ── 诊断：打印 Excel 读到的原名称（前10条）
            excel_keys = list(rename_map.keys())
            self._log(f"Excel原名称列示例（前{min(5,len(excel_keys))}条）:", 'info')
            for k in excel_keys[:5]:
                repr_k = repr(k)   # 显示原始字符，含隐藏字符
                self._log(f"  {repr_k}  →  {rename_map[k]['new_name']}", 'info')

            if mode == "wav":
                self._log("扫描WAV文件...", 'info')
                existing = self.get_all_wav_files(self.folder_path.get())
                self._log(f"共找到 {len(existing)} 个WAV文件", 'success')
                # 诊断：打印文件名示例
                ex_keys = list(existing.keys())
                self._log(f"文件夹WAV文件名示例（前{min(5,len(ex_keys))}条）:", 'info')
                for k in ex_keys[:5]:
                    self._log(f"  {repr(k)}", 'info')
            else:
                self._log("扫描第一层子文件夹...", 'info')
                existing = self.get_first_level_subfolders(self.folder_path.get())
                self._log(f"共找到 {len(existing)} 个子文件夹", 'success')
                # 诊断：打印子文件夹名示例
                ex_keys = list(existing.keys())
                self._log(f"子文件夹名称示例（前{min(10,len(ex_keys))}条）:", 'info')
                for k in ex_keys[:10]:
                    self._log(f"  {repr(k)}", 'info')

            preview_rows = []
            not_found = []
            # 文件夹模式：子字符串包含匹配；WAV模式：精确匹配 stem
            is_folder_mode = (mode == "folder")

            for old_name, info in rename_map.items():
                new_name = info['new_name']
                if is_folder_mode:
                    # 子字符串匹配：找出所有包含 old_name 的文件夹名
                    matched = [fn for fn in existing if old_name in fn]
                    if len(matched) == 0:
                        not_found.append(old_name)
                    elif len(matched) == 1:
                        folder_name = matched[0]
                        preview_rows.append({
                            'old': folder_name,
                            'new': new_name,
                            'old_path': existing[folder_name],
                            'row': info['row'],
                            'mode': mode,
                            'match_key': old_name,
                        })
                        self._log(f"匹配: [{old_name}]  ←  {folder_name}", 'info')
                    else:
                        # 多个文件夹包含该关键词，全部列入预览让用户确认
                        self._log(f"⚠ [{old_name}] 匹配到 {len(matched)} 个文件夹，全部加入预览请手动确认:", 'warning')
                        for folder_name in matched:
                            self._log(f"    - {folder_name}", 'warning')
                            preview_rows.append({
                                'old': folder_name,
                                'new': new_name,
                                'old_path': existing[folder_name],
                                'row': info['row'],
                                'mode': mode,
                                'match_key': old_name,
                            })
                else:
                    # WAV 模式：精确匹配 stem
                    if old_name in existing:
                        preview_rows.append({
                            'old': old_name,
                            'new': new_name,
                            'old_path': existing[old_name],
                            'row': info['row'],
                            'mode': mode,
                        })
                    else:
                        not_found.append(old_name)

            # 将预览数据加载到编辑器（必须在主线程执行）
            rows_copy = list(preview_rows)
            self.root.after(0, lambda: self.preview_editor.load_rows(rows_copy))
            self._log(f"匹配成功: {len(preview_rows)}/{len(rename_map)}", 'success')

            if not_found:
                self._log(f"未找到 {len(not_found)} 项:", 'warning')
                for name in not_found[:20]:
                    self._log(f"  - {repr(name)}", 'warning')
                if len(not_found) > 20:
                    self._log(f"  ... 等共 {len(not_found)} 项", 'warning')
                if not preview_rows and existing:
                    self._log("⚠ 全部未匹配！可能原因：", 'error')
                    self._log("  1. Excel列设置错误（检查原名称列是否为A列）", 'error')
                    self._log("  2. Excel名称与文件夹名大小写/空格不一致", 'error')
                    self._log("  3. 文件夹模式下请确保选择了【父级文件夹】", 'error')

            if preview_rows:
                self._log("预览完成！可在预览区修改/删除后点击【执行重命名】。", 'success')
            else:
                self._log("没有可预览的项目，请检查日志中的诊断信息。", 'warning')

        threading.Thread(target=_run, daemon=True).start()

    # ──────────────────────── 执行重命名（Tab1）────────────────────────
    def execute_rename(self):
        if self.preview_editor.is_empty():
            messagebox.showwarning("提示", "请先点击【预览重命名】生成预览！")
            return

        rows = self.preview_editor.get_rows()
        if not rows:
            messagebox.showwarning("提示", "预览区没有有效的重命名项（所有行已被删除或新名称为空）！")
            return

        if not messagebox.askyesno("确认执行",
                                   f"即将重命名 {len(rows)} 个项目，确认执行？"):
            return

        def _run():
            self._log(f"{'='*50}", 'info')
            self._log(f"开始执行重命名，共 {len(rows)} 项...", 'info')

            self.rename_results = []
            self.progress['maximum'] = len(rows)
            self.progress['value'] = 0

            # 按 mode 分组处理
            wav_rows = [r for r in rows if r.get('mode') == 'wav']
            folder_rows = [r for r in rows if r.get('mode') == 'folder']

            # 辅助：调度主线程更新预览区某行状态
            def _mark(pi, status, reason='', path=''):
                if pi is not None:
                    self.root.after(0, lambda: self.preview_editor.mark_result(
                        pi, status, reason, path))

            # --- WAV 文件重命名 ---
            if wav_rows:
                rename_map_subset = {r['old']: {'new_name': r['new'], 'row': r['row']}
                                     for r in wav_rows}
                existing_wav = {r['old']: r['old_path'] for r in wav_rows}
                _, conflicts = self.analyze_rename_dependencies(rename_map_subset, existing_wav)
                temp_mapping = {}  # {temp_path: (old_name, new_name, row, preview_idx, orig_path)}

                # 阶段1：冲突文件临时重命名
                if conflicts:
                    self._log(f"阶段1: 处理 {len(conflicts)} 个命名冲突...", 'info')
                    for r in wav_rows:
                        if r['old'] in conflicts:
                            pi = r.get('_preview_idx')
                            old_p = Path(r['old_path'])
                            temp_name = f"_tmp_{uuid.uuid4().hex[:8]}_.wav"
                            temp_p = old_p.parent / temp_name
                            try:
                                old_p.rename(temp_p)
                                temp_mapping[str(temp_p)] = (r['old'], r['new'], r['row'], pi, str(old_p))
                                self._log(f"临时: {r['old']}.wav -> {temp_name}", 'info')
                            except Exception as e:
                                self._log(f"临时重命名失败: {r['old']}.wav - {e}", 'error')
                                self.rename_results.append({
                                    'old_name': r['old'], 'new_name': r['new'],
                                    'old_path': str(old_p), 'new_path': '',
                                    'status': '失败', 'error': str(e), 'row': r['row']
                                })
                                _mark(pi, '失败', str(e), str(old_p))

                # 阶段2：普通重命名
                self._log("阶段2: 执行普通重命名...", 'info')
                for cnt, r in enumerate(wav_rows, 1):
                    if r['old'] in conflicts:
                        continue
                    pi = r.get('_preview_idx')
                    old_p = Path(r['old_path'])
                    new_p = old_p.parent / f"{r['new']}.wav"
                    self.progress['value'] = cnt
                    try:
                        if old_p == new_p:
                            self._log(f"跳过（同名）: {r['old']}.wav", 'info')
                            self.rename_results.append({
                                'old_name': r['old'], 'new_name': r['new'],
                                'old_path': str(old_p), 'new_path': str(new_p),
                                'status': '跳过', 'error': '名称相同', 'row': r['row']
                            })
                            _mark(pi, '跳过', '名称相同', str(old_p))
                            continue
                        if new_p.exists():
                            self._log(f"失败（目标已存在）: {r['new']}.wav", 'warning')
                            self.rename_results.append({
                                'old_name': r['old'], 'new_name': r['new'],
                                'old_path': str(old_p), 'new_path': str(new_p),
                                'status': '失败', 'error': '目标文件已存在', 'row': r['row']
                            })
                            _mark(pi, '失败', '目标文件已存在', str(old_p))
                            continue
                        old_p.rename(new_p)
                        self._log(f"✓ {r['old']}.wav  →  {r['new']}.wav", 'success')
                        self.rename_results.append({
                            'old_name': r['old'], 'new_name': r['new'],
                            'old_path': str(old_p), 'new_path': str(new_p),
                            'status': '成功', 'error': '', 'row': r['row']
                        })
                        _mark(pi, '成功', '', str(new_p))
                    except Exception as e:
                        self._log(f"✗ {r['old']}.wav - {e}", 'error')
                        self.rename_results.append({
                            'old_name': r['old'], 'new_name': r['new'],
                            'old_path': str(old_p), 'new_path': '',
                            'status': '失败', 'error': str(e), 'row': r['row']
                        })
                        _mark(pi, '失败', str(e), str(old_p))

                # 阶段3：临时文件最终重命名
                if temp_mapping:
                    self._log("阶段3: 完成循环依赖文件最终重命名...", 'info')
                    for temp_p_str, (old, new, row, pi, orig_path) in temp_mapping.items():
                        temp_p = Path(temp_p_str)
                        new_p = temp_p.parent / f"{new}.wav"
                        try:
                            temp_p.rename(new_p)
                            self._log(f"✓ {old}.wav  →  {new}.wav (循环)", 'success')
                            self.rename_results.append({
                                'old_name': old, 'new_name': new,
                                'old_path': orig_path,   # 使用原始路径，撤销时回到正确名称
                                'new_path': str(new_p),
                                'status': '成功', 'error': '', 'row': row
                            })
                            _mark(pi, '成功', '', str(new_p))
                        except Exception as e:
                            self._log(f"✗ {old}.wav - {e}", 'error')
                            self.rename_results.append({
                                'old_name': old, 'new_name': new,
                                'old_path': orig_path, 'new_path': '',
                                'status': '失败', 'error': str(e), 'row': row
                            })
                            _mark(pi, '失败', str(e), orig_path)

            # --- 文件夹重命名 ---
            if folder_rows:
                self._log("执行文件夹重命名...", 'info')
                for cnt, r in enumerate(folder_rows, 1):
                    pi = r.get('_preview_idx')
                    old_p = Path(r['old_path'])
                    new_p = old_p.parent / r['new']
                    self.progress['value'] = len(wav_rows) + cnt
                    try:
                        if old_p == new_p:
                            self._log(f"跳过（同名）: {r['old']}", 'info')
                            self.rename_results.append({
                                'old_name': r['old'], 'new_name': r['new'],
                                'old_path': str(old_p), 'new_path': str(new_p),
                                'status': '跳过', 'error': '名称相同', 'row': r['row']
                            })
                            _mark(pi, '跳过', '名称相同', str(old_p))
                            continue
                        if new_p.exists():
                            self._log(f"失败（目标已存在）: {r['new']}", 'warning')
                            self.rename_results.append({
                                'old_name': r['old'], 'new_name': r['new'],
                                'old_path': str(old_p), 'new_path': str(new_p),
                                'status': '失败', 'error': '目标文件夹已存在', 'row': r['row']
                            })
                            _mark(pi, '失败', '目标文件夹已存在', str(old_p))
                            continue
                        old_p.rename(new_p)
                        self._log(f"✓ {r['old']}  →  {r['new']}", 'success')
                        self.rename_results.append({
                            'old_name': r['old'], 'new_name': r['new'],
                            'old_path': str(old_p), 'new_path': str(new_p),
                            'status': '成功', 'error': '', 'row': r['row']
                        })
                        _mark(pi, '成功', '', str(new_p))
                    except Exception as e:
                        self._log(f"✗ {r['old']} - {e}", 'error')
                        self.rename_results.append({
                            'old_name': r['old'], 'new_name': r['new'],
                            'old_path': str(old_p), 'new_path': '',
                            'status': '失败', 'error': str(e), 'row': r['row']
                        })
                        _mark(pi, '失败', str(e), str(old_p))

            # 执行后不再回写 Excel（结果展示在预览区）
            # 保存撤销历史
            self._save_undo_history()

            # 确保进度条走满
            self.root.after(0, lambda: self.progress.configure(
                value=self.progress['maximum']))

            ok = sum(1 for r in self.rename_results if r['status'] == '成功')
            fail = sum(1 for r in self.rename_results if r['status'] == '失败')
            self._log(f"{'='*50}", 'info')
            self._log(f"重命名完成！  成功: {ok}  失败: {fail}", 'success' if fail == 0 else 'warning')
            if ok > 0:
                self._log("可使用【撤销重命名】按钮恢复本次操作", 'info')

        threading.Thread(target=_run, daemon=True).start()

    # ──────────────────────── 撤销 ────────────────────────
    def undo_rename(self):
        if not self.folder_path.get():
            messagebox.showwarning("提示", "请先选择文件夹！")
            return
        try:
            undo_dir = Path(self.folder_path.get()) / ".rename_undo"
            if not undo_dir.exists():
                messagebox.showinfo("提示", "没有可撤销的操作！")
                return
            undo_files = sorted(undo_dir.glob("undo_*.json"), reverse=True)
            if not undo_files:
                messagebox.showinfo("提示", "没有可撤销的操作！")
                return
            latest = undo_files[0]
            with open(latest, 'r', encoding='utf-8') as f:
                undo_data = json.load(f)
            ops = undo_data['operations']
            ts = undo_data['timestamp']
            if not messagebox.askyesno("确认撤销",
                                       f"撤销时间: {ts}\n操作数: {len(ops)} 个\n\n确认撤销？"):
                return
            self._log(f"{'='*50}", 'info')
            self._log(f"开始撤销 ({ts})...", 'info')
            ok = fail = 0

            # 撤销时同样可能存在循环依赖：
            # 撤销操作 = new_path -> old_path
            # 若某条操作的目标 old_path 已被其他文件占用，且那个文件也要被撤销，则需临时中转
            # 构建撤销时的 "new_stem -> old_stem" 映射表，与现存文件收集对比
            existing_names = set()
            for op in ops:
                np = Path(op['new_path'])
                if np.exists():
                    existing_names.add(np.stem)

            # 找出需要临时中转的项：目标名在现有文件中，且那个文件本身也要被撤销重命名
            old_stems_to_undo = {Path(op['new_path']).stem for op in ops if Path(op['new_path']).exists()}
            need_temp = set()
            for op in ops:
                old_stem = Path(op['old_path']).stem   # 撤销目标名
                new_stem = Path(op['new_path']).stem   # 当前文件名
                if old_stem in existing_names and old_stem in old_stems_to_undo:
                    need_temp.add(new_stem)             # 这个文件需要先临时

            undo_temp_map = {}  # {temp_path_str: (new_path_str, old_path_str)}

            # 阶段1：冲突项先临时重命名
            if need_temp:
                self._log(f"撤销小阶段1: 临时中转 {len(need_temp)} 个循环冲突文件...", 'info')
                for op in ops:
                    new_stem = Path(op['new_path']).stem
                    if new_stem in need_temp:
                        new_p = Path(op['new_path'])
                        if not new_p.exists():
                            continue
                        temp_name = f"_utmp_{uuid.uuid4().hex[:8]}_.wav"
                        temp_p = new_p.parent / temp_name
                        try:
                            new_p.rename(temp_p)
                            undo_temp_map[str(temp_p)] = (op['new_path'], op['old_path'])
                            self._log(f"临时: {new_p.name} -> {temp_name}", 'info')
                        except Exception as e:
                            self._log(f"临时重命名失败: {new_p.name} - {e}", 'error')
                            fail += 1

            # 阶段2：普通撤销（非循环冲突项）
            for idx, op in enumerate(ops, 1):
                new_stem = Path(op['new_path']).stem
                if new_stem in need_temp:
                    continue  # 已在阶段1处理
                try:
                    new_p = Path(op['new_path'])
                    old_p = Path(op['old_path'])
                    if not new_p.exists():
                        self._log(f"警告: 文件不存在 {new_p.name}", 'warning')
                        fail += 1
                        continue
                    new_p.rename(old_p)
                    self._log(f"撤销[{idx}/{len(ops)}]: {new_p.name} -> {old_p.name}", 'success')
                    ok += 1
                except Exception as e:
                    self._log(f"撤销失败[{idx}]: {e}", 'error')
                    fail += 1

            # 阶段3：临时文件最终重命名
            if undo_temp_map:
                self._log("撤销小阶段3: 完成循环冲突文件最终重命名...", 'info')
                for temp_p_str, (new_path_str, old_path_str) in undo_temp_map.items():
                    temp_p = Path(temp_p_str)
                    old_p = Path(old_path_str)
                    try:
                        temp_p.rename(old_p)
                        self._log(f"撤销(循环): {Path(new_path_str).name} -> {old_p.name}", 'success')
                        ok += 1
                    except Exception as e:
                        self._log(f"撤销失败(循环): {old_p.name} - {e}", 'error')
                        fail += 1
            if self.excel_path.get() == undo_data.get('excel_file'):
                self._clear_excel_status(ops)
            latest.unlink()
            self._log(f"撤销完成！成功: {ok}, 失败: {fail}", 'success')
            messagebox.showinfo("完成", f"撤销完成！\n✅ 成功: {ok}\n❌ 失败: {fail}")
        except Exception as e:
            self._log(f"撤销失败: {e}", 'error')
            messagebox.showerror("错误", f"撤销失败:\n{e}")

    def _save_undo_history(self):
        try:
            ops = [r for r in self.rename_results if r['status'] == '成功']
            if not ops:
                return
            undo_dir = Path(self.folder_path.get()) / ".rename_undo"
            undo_dir.mkdir(exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            undo_file = undo_dir / f"undo_{ts}.json"
            with open(undo_file, 'w', encoding='utf-8') as f:
                json.dump({'timestamp': ts,
                           'excel_file': self.excel_path.get(),
                           'operations': ops}, f, ensure_ascii=False, indent=2)
            self._log(f"撤销历史已保存: {undo_file.name}", 'info')
        except Exception as e:
            self._log(f"保存撤销历史失败: {e}", 'warning')

    def _clear_excel_status(self, operations):
        """不再操作Excel状态列，此方法保留但不执行任何操作"""
        pass

    # ──────────────────────── Tab2: 删除字符串重命名 WAV ────────────────────────
    def _strip_preview(self):
        folder = self.strip_folder_path.get()
        strip_str = self.strip_string.get().strip()
        if not folder:
            messagebox.showwarning("提示", "请先选择文件夹！")
            return
        if not strip_str:
            messagebox.showwarning("提示", "请输入要删除的字符串！")
            return

        def _run():
            self._strip_log(f"{'='*50}", 'info')
            self._strip_log(f"扫描WAV文件（要删除字符串: 「{strip_str}」）...", 'info')
            folder_p = Path(folder)
            wav_files = list(folder_p.rglob("*.wav"))
            self._strip_log(f"共找到 {len(wav_files)} 个WAV文件", 'success')

            rows = []
            skipped = []
            for wav in wav_files:
                stem = wav.stem
                if strip_str in stem:
                    new_stem = stem.replace(strip_str, '')
                    if new_stem:
                        rows.append({
                            'old': stem,
                            'new': new_stem,
                            'old_path': str(wav),
                            'parent': str(wav.parent),
                        })
                    else:
                        skipped.append(f"{stem}.wav（删除后名称为空，已跳过）")
                # 不含目标字符串的文件不显示

            self.root.after(0, lambda: self.strip_preview_editor.load_rows(rows))
            self._strip_log(f"匹配到 {len(rows)} 个需要修改的文件", 'success')
            if skipped:
                self._strip_log(f"跳过 {len(skipped)} 个（删除后名称为空）:", 'warning')
                for s in skipped:
                    self._strip_log(f"  - {s}", 'warning')
            if not rows:
                self._strip_log("没有找到包含该字符串的文件名。", 'warning')
            else:
                self._strip_log("预览完成！可在预览区修改/删除后点击【执行重命名】。", 'success')

        threading.Thread(target=_run, daemon=True).start()

    def _strip_execute(self):
        if self.strip_preview_editor.is_empty():
            messagebox.showwarning("提示", "请先点击【预览重命名】生成预览！")
            return
        rows = self.strip_preview_editor.get_rows()
        if not rows:
            messagebox.showwarning("提示", "预览区没有有效的重命名项！")
            return
        if not messagebox.askyesno("确认执行",
                                   f"即将重命名 {len(rows)} 个WAV文件，确认执行？"):
            return

        def _run():
            self._strip_log(f"{'='*50}", 'info')
            self._strip_log(f"开始执行，共 {len(rows)} 项...", 'info')
            self.strip_progress['maximum'] = len(rows)
            self.strip_progress['value'] = 0
            ok = fail = skip = 0
            strip_ops = []   # 记录成功操作，用于撤销

            def _smark(pi, status, reason='', path=''):
                if pi is not None:
                    self.root.after(0, lambda: self.strip_preview_editor.mark_result(
                        pi, status, reason, path))

            for idx, r in enumerate(rows, 1):
                pi = r.get('_preview_idx')
                self.strip_progress['value'] = idx
                old_p = Path(r['old_path'])
                new_p = old_p.parent / f"{r['new']}.wav"
                try:
                    if old_p == new_p:
                        self._strip_log(f"跳过（同名）: {r['old']}.wav", 'info')
                        skip += 1
                        _smark(pi, '跳过', '名称相同', str(old_p))
                        continue
                    if new_p.exists():
                        self._strip_log(f"失败（目标已存在）: {r['new']}.wav", 'warning')
                        fail += 1
                        _smark(pi, '失败', '目标文件已存在', str(old_p))
                        continue
                    old_p.rename(new_p)
                    self._strip_log(f"✓ {r['old']}.wav  →  {r['new']}.wav", 'success')
                    ok += 1
                    strip_ops.append({
                        'old_name': r['old'],
                        'new_name': r['new'],
                        'old_path': str(old_p),
                        'new_path': str(new_p),
                    })
                    _smark(pi, '成功', '', str(new_p))
                except Exception as e:
                    self._strip_log(f"✗ {r['old']}.wav - {e}", 'error')
                    fail += 1
                    _smark(pi, '失败', str(e), str(old_p))

            # 保存撤销历史
            if strip_ops:
                self._strip_save_undo(strip_ops)

            self._strip_log(f"{'='*50}", 'info')
            self._strip_log(
                f"完成！  成功: {ok}  失败: {fail}  跳过: {skip}",
                'success' if fail == 0 else 'warning'
            )
            if ok > 0:
                self._strip_log("可使用【撤销重命名】按钮恢复本次操作", 'info')

        threading.Thread(target=_run, daemon=True).start()

    def _strip_save_undo(self, ops: List[Dict]):
        """保存 Tab2 撤销历史到 strip_folder/.rename_undo/"""
        try:
            folder = self.strip_folder_path.get()
            if not folder:
                return
            undo_dir = Path(folder) / ".rename_undo"
            undo_dir.mkdir(exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            undo_file = undo_dir / f"strip_undo_{ts}.json"
            with open(undo_file, 'w', encoding='utf-8') as f:
                json.dump({'timestamp': ts, 'operations': ops},
                          f, ensure_ascii=False, indent=2)
            self._strip_log(f"撤销历史已保存: {undo_file.name}", 'info')
        except Exception as e:
            self._strip_log(f"保存撤销历史失败: {e}", 'warning')

    def _strip_undo(self):
        """撤销 Tab2 最近一次删除字符串重命名"""
        folder = self.strip_folder_path.get()
        if not folder:
            messagebox.showwarning("提示", "请先选择文件夹！")
            return
        try:
            undo_dir = Path(folder) / ".rename_undo"
            if not undo_dir.exists():
                messagebox.showinfo("提示", "没有可撤销的操作！")
                return
            # 只查找 strip_undo_*.json
            undo_files = sorted(undo_dir.glob("strip_undo_*.json"), reverse=True)
            if not undo_files:
                messagebox.showinfo("提示", "没有可撤销的删除字符串重命名操作！")
                return
            latest = undo_files[0]
            with open(latest, 'r', encoding='utf-8') as f:
                undo_data = json.load(f)
            ops = undo_data['operations']
            ts = undo_data['timestamp']
            if not messagebox.askyesno("确认撤销",
                                       f"撤销时间: {ts}\n操作数: {len(ops)} 个\n\n确认撤销？"):
                return
            self._strip_log(f"{'='*50}", 'info')
            self._strip_log(f"开始撤销 ({ts})...", 'info')
            ok = fail = 0
            for idx, op in enumerate(ops, 1):
                try:
                    new_p = Path(op['new_path'])
                    old_p = Path(op['old_path'])
                    if not new_p.exists():
                        self._strip_log(f"警告: 文件不存在 {new_p.name}", 'warning')
                        fail += 1
                        continue
                    new_p.rename(old_p)
                    self._strip_log(
                        f"撤销[{idx}/{len(ops)}]: {new_p.name} -> {old_p.name}", 'success')
                    ok += 1
                except Exception as e:
                    self._strip_log(f"撤销失败[{idx}]: {e}", 'error')
                    fail += 1
            latest.unlink()
            self._strip_log(f"撤销完成！成功: {ok}, 失败: {fail}", 'success')
            messagebox.showinfo("完成", f"撤销完成！\n✅ 成功: {ok}\n❌ 失败: {fail}")
        except Exception as e:
            self._strip_log(f"撤销失败: {e}", 'error')
            messagebox.showerror("错误", f"撤销失败:\n{e}")


# ─────────────────────────────────────────────
def main():
    root = tk.Tk()
    app = WavRenamerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
